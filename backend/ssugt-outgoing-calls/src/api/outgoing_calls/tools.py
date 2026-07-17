import re
from io import BytesIO
from typing import Optional, List, Dict, Any
from urllib.parse import urlparse
from zoneinfo import ZoneInfo

import openpyxl
from dateutil import parser
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter

XLSX_MAX_TITLE_LENGTH = 30
KZ_SECOND_DIGITS = {"0", "1", "5", "6", "7"}
AB_OR_SO_SECOND_DIGITS = {"929", "940"}


def build_amqp_url(*, host: str, port: int, user: str, password: str) -> str:
    return f"amqp://{user}:{password}@{host}:{port}/"


def extract_host_port(
    url: str, default_port: Optional[int] = None
) -> tuple[str, Optional[int]]:
    parsed = urlparse(url)

    host = parsed.hostname
    port = parsed.port

    if port is None:
        if parsed.scheme == "http":
            port = 80
        elif parsed.scheme == "https":
            port = 443
        elif default_port is not None:
            port = default_port

    return host, port


def get_phone_category(digits):
    if len(digits) == 11 and digits.startswith("7"):
        second = digits[1]
        triple = digits[1:4]

        # Казахстан
        if second in KZ_SECOND_DIGITS:
            return "international_numbers"

        # Абхазия/Южная Осетия
        if triple in AB_OR_SO_SECOND_DIGITS:
            return "international_numbers"

        # Россия
        if second == "9":
            return "ru_mobile_numbers"

        # Городские
        return "ru_city_numbers"

    return "international_numbers"


def normalize_phone(text):
    match = re.search(r"(\+?\d[\d\s\-\(\)]{6,}\d)", text)
    if not match:
        return None

    raw_phone = match.group()

    digits = re.sub(r"\D", "", raw_phone)

    if digits.startswith("8") and len(digits) == 11:
        digits = "7" + digits[1:]

    elif digits.startswith("9") and len(digits) == 10:
        digits = "7" + digits

    if not 7 <= len(digits) <= 15:
        return None

    return digits


def reformat_datetime_text(value: Optional[str]) -> str:
    if value is None:
        return ""

    try:
        dt_utc = parser.parse(value)
        if dt_utc.tzinfo is None:
            dt_utc = dt_utc.replace(tzinfo=ZoneInfo("UTC"))
        else:
            dt_utc = dt_utc.astimezone(ZoneInfo("UTC"))

        dt_local = dt_utc.astimezone(ZoneInfo("Asia/Novosibirsk"))

        return dt_local.strftime("%d.%m.%y %H:%M")
    except Exception:
        return ""


def reformat_status_text(value: str) -> str:
    if value == "done":
        return "Завершено"
    elif value == "ringing":
        return "Идет вызов"
    elif value == "failed":
        return "Неудачно"
    elif value == "in_queue":
        return "В очереди"

    return ""


def reformat_column_text(value: Any) -> str:
    if value is None:
        return ""

    return str(value)


def xlsx_clean_value(value):
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub("", value)
    return value


def json_to_xlsx(
    *,
    title: str,
    data: List[Dict[str, Any]],
    columns: Dict[str, str],
    highlight: Optional[Dict[str, Dict[str, str]]] = None,
    fixed_column_width: Optional[int] = None,
    wrap_text: bool = True,
) -> BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:XLSX_MAX_TITLE_LENGTH]

    headers_eng = list(columns.keys())

    headers = list(columns.values()) if columns else []
    ws.append(headers)

    for index, x in enumerate(data):
        items = [x.get(col) for col in headers_eng]
        items = [xlsx_clean_value(item) for item in items]

        ws.append(items)

        row_index = index + 2

        for col_name in headers_eng:
            col_index = headers_eng.index(col_name) + 1
            cell = ws.cell(row=row_index, column=col_index)

            if wrap_text:
                cell.alignment = Alignment(wrap_text=True)

        if highlight:
            for col_name, color_map in highlight.items():
                if col_name not in headers_eng:
                    continue

                value = str(x.get(col_name, ""))
                color_code = color_map.get(value)

                if not color_code:
                    continue

                col_index = headers_eng.index(col_name) + 1
                cell = ws.cell(row=row_index, column=col_index)
                cell.fill = PatternFill(
                    start_color=color_code, end_color=color_code, fill_type="solid"
                )

    for i, col_key in enumerate(headers_eng, 1):
        max_len = len(headers[i - 1])

        for x in data:
            value = x.get(col_key)

            if value is None:
                continue

            max_len = max(max_len, len(str(value)))

        width = max_len + 10

        if fixed_column_width:
            width = min(max_len + 10, fixed_column_width)

        ws.column_dimensions[get_column_letter(i)].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output
