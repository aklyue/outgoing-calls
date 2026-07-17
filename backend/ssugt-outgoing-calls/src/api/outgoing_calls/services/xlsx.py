import subprocess
import tempfile
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import List, Any, Optional
from collections import OrderedDict

import pandas as pd
from openpyxl import load_workbook
from pandas import ExcelWriter


@dataclass
class ConcatColumnConfig:
    concated_column: str
    columns: List[str]
    priority_column: Optional[str] = None


def convert_to_xlsx(file_bytes: bytes, filename: str) -> bytes:
    with tempfile.TemporaryDirectory() as tmpdir:
        tmpdir_path = Path(tmpdir)
        input_path = tmpdir_path / filename

        with open(input_path, "wb") as f:
            f.write(file_bytes)

        result = subprocess.run(
            [
                "libreoffice",
                "--headless",
                "--convert-to",
                "xlsx",
                "--outdir",
                str(tmpdir_path),
                str(input_path),
            ],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
        )

        if result.returncode != 0:
            raise RuntimeError(f"LibreOffice error: {result.stderr.decode()}")

        output_path = tmpdir_path / (input_path.stem + ".xlsx")
        if not output_path.exists():
            raise FileNotFoundError("Файл не сконвертирован.")

        with open(output_path, "rb") as f:
            return f.read()


def merge_tables(xlsx_bytes: bytes, concat_columns: List[ConcatColumnConfig]) -> bytes:
    wb = load_workbook(filename=BytesIO(xlsx_bytes), data_only=True)
    sheet = wb.active

    def is_cell_has_border(cell: Any) -> bool:
        has_border = cell.border and any(
            [
                cell.border.left.style,
                cell.border.right.style,
                cell.border.top.style,
                cell.border.bottom.style,
            ]
        )
        return has_border

    tables: List[List[List[Any]]] = []
    current_table: List[List[Any]] = []

    for row in sheet.iter_rows():
        row_data = []
        significant = False
        for cell in row:
            if is_cell_has_border(cell):
                significant = True
            row_data.append(cell.value)
        if significant:
            current_table.append(row_data)
        else:
            if current_table:
                tables.append(current_table)
                current_table = []
    if current_table:
        tables.append(current_table)

    def normalize_header(header_row: List[Any]) -> List[Optional[str]]:
        seen = {}
        result = []
        for col in header_row:
            name = str(col).strip() if col else ""
            if not name:
                result.append(None)
                continue
            original_name = name
            while name in seen:
                seen[original_name] += 1
                name = f"{original_name}_{seen[original_name]}"
            seen[name] = 1
            result.append(name)
        return result

    dfs: List[pd.DataFrame] = []
    column_set: OrderedDict[str, None] = OrderedDict()

    for table in tables:
        if not table:
            continue
        header = normalize_header(table[0])

        rows = table[1:]

        valid_indices = [i for i, name in enumerate(header) if name]
        filtered_header = [header[i] for i in valid_indices]

        cleaned_data = [
            [row[i] if i < len(row) else None for i in valid_indices] for row in rows
        ]

        df = pd.DataFrame(cleaned_data, columns=filtered_header)

        for config in concat_columns:
            new_col = config.concated_column
            cols_to_merge = config.columns
            priority = config.priority_column or cols_to_merge[0]

            def combine_row(row: pd.Series) -> Any:
                values = [row.get(col) for col in cols_to_merge if row.get(col)]
                if not values:
                    return ""
                if all(row.get(col) for col in cols_to_merge):
                    return row.get(priority)
                return values[0]

            df[new_col] = df.apply(combine_row, axis=1)

            for col in cols_to_merge:
                if col in df.columns and col != new_col:
                    df.drop(columns=col, inplace=True)

        for col in df.columns:
            column_set[col] = None
        dfs.append(df)

    final_columns = list(column_set.keys())
    final_dfs = [df.reindex(columns=final_columns) for df in dfs]

    final_dfs = [
        df.dropna(axis=1, how="all")
        for df in final_dfs
        if not df.empty and not df.isna().all(axis=None)
    ]

    if not final_dfs:
        return xlsx_bytes

    combined_df = pd.concat(final_dfs, ignore_index=True)

    with tempfile.NamedTemporaryFile(suffix=".xlsx") as tmp:
        with ExcelWriter(tmp.name, engine="openpyxl") as writer:
            combined_df.to_excel(writer, index=False)

        with open(tmp.name, "rb") as f:
            output_bytes = f.read()

    return output_bytes
